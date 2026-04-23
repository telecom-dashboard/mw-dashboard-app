from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import asc, desc, func, or_
from sqlalchemy.orm import Session, aliased

from app.core.database import get_db
from app.models.site_connectivity import SiteConnectivity
from app.models.microwave_link_budget import MicrowaveLinkBudget
from app.models.user import User
from app.routers.auth import require_admin
from app.utils.audit import create_audit_log, model_to_audit_dict

router = APIRouter(prefix="/site-connectivity", tags=["Site Connectivity"])


ALLOWED_SORT_FIELDS = {
    "id": SiteConnectivity.id,
    "sitea_id": SiteConnectivity.sitea_id,
    "siteb_id": SiteConnectivity.siteb_id,
    "link_id": SiteConnectivity.link_id,
    "category_ne": SiteConnectivity.category_ne,
    "depth": SiteConnectivity.depth,
    "dependency": SiteConnectivity.dependency,
    "pop_site": SiteConnectivity.pop_site,
    "child_site_connectivity": SiteConnectivity.child_site_connectivity,
    "child_site_name": SiteConnectivity.child_site_name,
    "is_active": SiteConnectivity.is_active,
    "created_at": SiteConnectivity.created_at,
    "updated_at": SiteConnectivity.updated_at,
}

TEMPLATE_HEADERS = [
    "SiteA ID",
    "SiteB ID",
    "Link ID",
    "Category [NE]",
    "Depth",
    "Dependency",
    "POP Site",
    "Child Site connectivity",
    "Child Site Name",
]


def clean_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def clean_int(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def get_attr(obj: Any, *names: str) -> Any:
    if obj is None:
        return None
    for name in names:
        if hasattr(obj, name):
            return getattr(obj, name)
    return None


def serialize_row(
    site_item: SiteConnectivity,
    budget_item: MicrowaveLinkBudget | None,
) -> dict[str, Any]:
    return {
        "id": site_item.id,
        "sitea_id": site_item.sitea_id,
        "siteb_id": site_item.siteb_id,
        "link_id": site_item.link_id,
        "category_ne": site_item.category_ne,
        "depth": site_item.depth,
        "dependency": site_item.dependency,
        "pop_site": site_item.pop_site,
        "child_site_connectivity": site_item.child_site_connectivity,
        "child_site_name": site_item.child_site_name,
        "is_active": site_item.is_active,
        "created_at": site_item.created_at.isoformat() if site_item.created_at else None,
        "updated_at": site_item.updated_at.isoformat() if site_item.updated_at else None,
        "budget_vendor": get_attr(budget_item, "vendor"),
        "budget_site_name_s1": get_attr(budget_item, "site_name_s1"),
        "budget_site_name_s2": get_attr(budget_item, "site_name_s2"),
        "budget_region": get_attr(budget_item, "region", "state_province"),
        "budget_township": get_attr(budget_item, "township"),
        "budget_zone": get_attr(budget_item, "zone"),
        "budget_ring_id_span_name": get_attr(budget_item, "ring_id_span_name"),
        "budget_media_type": get_attr(budget_item, "media_type"),
        "budget_revise": get_attr(budget_item, "revise"),
        "budget_site_name_s1_ip": get_attr(budget_item, "site_name_s1_ip"),
        "budget_site_name_s2_ip": get_attr(budget_item, "site_name_s2_ip"),
        "budget_site_name_s1_port": get_attr(budget_item, "site_name_s1_port"),
        "budget_site_name_s2_port": get_attr(budget_item, "site_name_s2_port"),
        "budget_link_class": get_attr(budget_item, "link_class"),
        "budget_model": get_attr(budget_item, "model"),
        "budget_status": get_attr(budget_item, "status"),
        "budget_protocol": get_attr(budget_item, "protocol"),
        "budget_comment": get_attr(budget_item, "comment"),
        "budget_type": get_attr(budget_item, "type"),
        "budget_bandwidth": get_attr(budget_item, "bandwidth"),
        "budget_planning_capacity": get_attr(budget_item, "planning_capacity"),
    }


def build_base_query(
    db: Session,
    search: str = "",
    category: str = "",
):
    budget_alias = aliased(MicrowaveLinkBudget)

    query = (
        db.query(SiteConnectivity, budget_alias)
        .outerjoin(
            budget_alias,
            or_(
                SiteConnectivity.link_id == budget_alias.link_id,
                SiteConnectivity.link_id == budget_alias.revise,
            ),
        )
    )

    if search:
        like = f"%{search.strip()}%"
        query = query.filter(
            or_(
                SiteConnectivity.sitea_id.ilike(like),
                SiteConnectivity.siteb_id.ilike(like),
                SiteConnectivity.link_id.ilike(like),
                SiteConnectivity.category_ne.ilike(like),
                SiteConnectivity.dependency.ilike(like),
                SiteConnectivity.pop_site.ilike(like),
                SiteConnectivity.child_site_connectivity.ilike(like),
                SiteConnectivity.child_site_name.ilike(like),
                budget_alias.vendor.ilike(like),
                budget_alias.site_name_s1.ilike(like),
                budget_alias.site_name_s2.ilike(like),
                budget_alias.site_name_s1_ip.ilike(like),
                budget_alias.site_name_s2_ip.ilike(like),
                budget_alias.model.ilike(like),
                budget_alias.status.ilike(like),
                budget_alias.protocol.ilike(like),
            )
        )

    if category and category.strip().lower() != "all":
        query = query.filter(SiteConnectivity.category_ne == category.strip())

    return query, budget_alias


@router.get("/category-options")
def get_site_connectivity_category_options(db: Session = Depends(get_db)):
    rows = (
        db.query(SiteConnectivity.category_ne)
        .filter(
            SiteConnectivity.category_ne.isnot(None),
            SiteConnectivity.category_ne != "",
        )
        .distinct()
        .order_by(asc(SiteConnectivity.category_ne))
        .all()
    )

    return {"items": [row[0] for row in rows if row[0]]}


@router.get("/summary")
def site_connectivity_summary(
    category: str = Query(default=""),
    db: Session = Depends(get_db),
):
    base_query = db.query(SiteConnectivity)

    if category and category.strip().lower() != "all":
        base_query = base_query.filter(SiteConnectivity.category_ne == category.strip())

    total = base_query.with_entities(func.count(SiteConnectivity.id)).scalar() or 0
    active = (
        base_query.filter(SiteConnectivity.is_active.is_(True))
        .with_entities(func.count(SiteConnectivity.id))
        .scalar()
        or 0
    )
    inactive = (
        base_query.filter(SiteConnectivity.is_active.is_(False))
        .with_entities(func.count(SiteConnectivity.id))
        .scalar()
        or 0
    )
    with_link = (
        base_query.filter(
            SiteConnectivity.link_id.isnot(None),
            SiteConnectivity.link_id != "",
        )
        .with_entities(func.count(SiteConnectivity.id))
        .scalar()
        or 0
    )

    return {
        "total_records": total,
        "active_records": active,
        "inactive_records": inactive,
        "with_link_id": with_link,
    }


@router.get("/")
def list_site_connectivity(
    search: str = Query(default=""),
    category: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=500),
    sort_by: str = Query(default="id"),
    sort_order: str = Query(default="desc"),
    db: Session = Depends(get_db),
):
    query, budget_alias = build_base_query(db, search, category)

    total_query = (
        db.query(SiteConnectivity)
        .outerjoin(
            budget_alias,
            or_(
                SiteConnectivity.link_id == budget_alias.link_id,
                SiteConnectivity.link_id == budget_alias.revise,
            ),
        )
    )

    if search:
        like = f"%{search.strip()}%"
        total_query = total_query.filter(
            or_(
                SiteConnectivity.sitea_id.ilike(like),
                SiteConnectivity.siteb_id.ilike(like),
                SiteConnectivity.link_id.ilike(like),
                SiteConnectivity.category_ne.ilike(like),
                SiteConnectivity.dependency.ilike(like),
                SiteConnectivity.pop_site.ilike(like),
                SiteConnectivity.child_site_connectivity.ilike(like),
                SiteConnectivity.child_site_name.ilike(like),
                budget_alias.vendor.ilike(like),
                budget_alias.site_name_s1.ilike(like),
                budget_alias.site_name_s2.ilike(like),
                budget_alias.site_name_s1_ip.ilike(like),
                budget_alias.site_name_s2_ip.ilike(like),
                budget_alias.model.ilike(like),
                budget_alias.status.ilike(like),
                budget_alias.protocol.ilike(like),
            )
        )

    if category and category.strip().lower() != "all":
        total_query = total_query.filter(SiteConnectivity.category_ne == category.strip())

    total = total_query.with_entities(func.count(func.distinct(SiteConnectivity.id))).scalar() or 0

    sort_column = ALLOWED_SORT_FIELDS.get(sort_by, SiteConnectivity.id)
    ordering = desc(sort_column) if sort_order.lower() == "desc" else asc(sort_column)

    rows = (
        query.order_by(ordering)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = [serialize_row(site_item, budget_item) for site_item, budget_item in rows]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if page_size else 1,
    }


@router.post("/")
def create_site_connectivity(
    payload: dict[str, Any],
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    item = SiteConnectivity(
        sitea_id=clean_str(payload.get("sitea_id")),
        siteb_id=clean_str(payload.get("siteb_id")),
        link_id=clean_str(payload.get("link_id")),
        category_ne=clean_str(payload.get("category_ne")),
        depth=clean_int(payload.get("depth")),
        dependency=clean_str(payload.get("dependency")),
        pop_site=clean_str(payload.get("pop_site")),
        child_site_connectivity=clean_str(payload.get("child_site_connectivity")),
        child_site_name=clean_str(payload.get("child_site_name")),
        is_active=bool(payload.get("is_active", True)),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(item)
    db.flush()
    create_audit_log(
        db,
        table_name="site_connectivity",
        record_id=item.id,
        action="create",
        current_user=current_user,
        new_values=model_to_audit_dict(item),
    )
    db.commit()
    db.refresh(item)
    return {"message": "Site connectivity record created successfully", "id": item.id}


@router.put("/{item_id}")
def update_site_connectivity(
    item_id: int,
    payload: dict[str, Any],
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    item = db.query(SiteConnectivity).filter(SiteConnectivity.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Site connectivity record not found")

    old_values = model_to_audit_dict(item)
    item.sitea_id = clean_str(payload.get("sitea_id"))
    item.siteb_id = clean_str(payload.get("siteb_id"))
    item.link_id = clean_str(payload.get("link_id"))
    item.category_ne = clean_str(payload.get("category_ne"))
    item.depth = clean_int(payload.get("depth"))
    item.dependency = clean_str(payload.get("dependency"))
    item.pop_site = clean_str(payload.get("pop_site"))
    item.child_site_connectivity = clean_str(payload.get("child_site_connectivity"))
    item.child_site_name = clean_str(payload.get("child_site_name"))
    item.is_active = bool(payload.get("is_active", True))
    item.updated_at = datetime.utcnow()

    create_audit_log(
        db,
        table_name="site_connectivity",
        record_id=item.id,
        action="update",
        current_user=current_user,
        old_values=old_values,
        new_values=model_to_audit_dict(item),
    )
    db.commit()
    db.refresh(item)
    return {"message": "Site connectivity record updated successfully"}


@router.delete("/{item_id}")
def delete_site_connectivity(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    item = db.query(SiteConnectivity).filter(SiteConnectivity.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Site connectivity record not found")

    old_values = model_to_audit_dict(item)
    db.delete(item)
    create_audit_log(
        db,
        table_name="site_connectivity",
        record_id=item_id,
        action="delete",
        current_user=current_user,
        old_values=old_values,
    )
    db.commit()
    return {"message": "Site connectivity record deleted successfully"}


@router.post("/bulk-delete")
def bulk_delete_site_connectivity(
    payload: dict[str, Any],
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    ids = payload.get("ids") or []
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")

    items = (
        db.query(SiteConnectivity)
        .filter(SiteConnectivity.id.in_(ids))
        .all()
    )
    old_values = [model_to_audit_dict(item) for item in items]
    db.query(SiteConnectivity).filter(SiteConnectivity.id.in_(ids)).delete(synchronize_session=False)
    create_audit_log(
        db,
        table_name="site_connectivity",
        record_id=0,
        action="bulk_delete",
        current_user=current_user,
        old_values={"items": old_values},
        new_values={"deleted_ids": ids, "deleted_count": len(items)},
    )
    db.commit()
    return {"message": "Selected site connectivity records deleted successfully"}


@router.delete("/")
def delete_all_site_connectivity(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    deleted_count = db.query(func.count(SiteConnectivity.id)).scalar() or 0
    db.query(SiteConnectivity).delete(synchronize_session=False)
    create_audit_log(
        db,
        table_name="site_connectivity",
        record_id=0,
        action="delete_all",
        current_user=current_user,
        new_values={"deleted_count": deleted_count},
    )
    db.commit()
    return {"message": "All site connectivity records deleted successfully"}


@router.post("/import")
async def import_site_connectivity_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")

    content = await file.read()
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_index = {header: idx for idx, header in enumerate(headers)}

    missing = [col for col in TEMPLATE_HEADERS if col not in header_index]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required Excel columns: {', '.join(missing)}",
        )

    created = 0

    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        item = SiteConnectivity(
            sitea_id=clean_str(row[header_index["SiteA ID"]]),
            siteb_id=clean_str(row[header_index["SiteB ID"]]),
            link_id=clean_str(row[header_index["Link ID"]]),
            category_ne=clean_str(row[header_index["Category [NE]"]]),
            depth=clean_int(row[header_index["Depth"]]),
            dependency=clean_str(row[header_index["Dependency"]]),
            pop_site=clean_str(row[header_index["POP Site"]]),
            child_site_connectivity=clean_str(row[header_index["Child Site connectivity"]]),
            child_site_name=clean_str(row[header_index["Child Site Name"]]),
            is_active=True,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(item)
        created += 1

    db.commit()
    create_audit_log(
        db,
        table_name="site_connectivity",
        record_id=0,
        action="import",
        current_user=current_user,
        new_values={
            "filename": file.filename,
            "created": created,
            "updated": 0,
        },
    )
    db.commit()

    return {
        "message": "Excel import completed successfully",
        "created": created,
        "updated": 0,
        "errors": [],
    }


@router.get("/export/template")
def export_site_connectivity_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Site Connectivity Template"

    sheet.append(TEMPLATE_HEADERS)
    sheet.append(
        [
            "SITE001",
            "SITE002",
            "AWD0001-AWD0002",
            "Access",
            1,
            "Primary",
            "POP-YGN",
            "Connected",
            "Child Site A",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=site_connectivity_template.xlsx"},
    )


@router.get("/export/excel")
def export_site_connectivity_excel(
    search: str = Query(default=""),
    category: str = Query(default=""),
    db: Session = Depends(get_db),
):
    query, _ = build_base_query(db, search, category)
    rows = query.order_by(desc(SiteConnectivity.id)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Site Connectivity"

    headers = [
        "ID",
        "SiteA ID",
        "SiteB ID",
        "Link ID",
        "Category [NE]",
        "Depth",
        "Dependency",
        "POP Site",
        "Child Site connectivity",
        "Child Site Name",
        "Active",
        "Vendor",
        "Site Name S1",
        "Site Name S2",
        "Revise",
        "Protocol",
        "Status",
        "Model",
        "Bandwidth",
        "Planning Capacity",
        "Site S1 IP",
        "Site S2 IP",
        "Site S1 Port",
        "Site S2 Port",
    ]
    sheet.append(headers)

    for site_item, budget_item in rows:
        row = serialize_row(site_item, budget_item)
        sheet.append(
            [
                row["id"],
                row["sitea_id"],
                row["siteb_id"],
                row["link_id"],
                row["category_ne"],
                row["depth"],
                row["dependency"],
                row["pop_site"],
                row["child_site_connectivity"],
                row["child_site_name"],
                "Active" if row["is_active"] else "Inactive",
                row["budget_vendor"],
                row["budget_site_name_s1"],
                row["budget_site_name_s2"],
                row["budget_revise"],
                row["budget_protocol"],
                row["budget_status"],
                row["budget_model"],
                row["budget_bandwidth"],
                row["budget_planning_capacity"],
                row["budget_site_name_s1_ip"],
                row["budget_site_name_s2_ip"],
                row["budget_site_name_s1_port"],
                row["budget_site_name_s2_port"],
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=site_connectivity.xlsx"},
    )


@router.post("/export/selected-excel")
def export_selected_site_connectivity_excel(
    payload: dict[str, Any],
    db: Session = Depends(get_db),
):
    ids = payload.get("ids") or []
    if not ids:
        raise HTTPException(status_code=400, detail="No selected IDs provided")

    budget_alias = aliased(MicrowaveLinkBudget)

    rows = (
        db.query(SiteConnectivity, budget_alias)
        .outerjoin(
            budget_alias,
            or_(
                SiteConnectivity.link_id == budget_alias.link_id,
                SiteConnectivity.link_id == budget_alias.revise,
            ),
        )
        .filter(SiteConnectivity.id.in_(ids))
        .order_by(desc(SiteConnectivity.id))
        .all()
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Selected Site Connectivity"

    headers = [
        "ID",
        "SiteA ID",
        "SiteB ID",
        "Link ID",
        "Category [NE]",
        "Depth",
        "Dependency",
        "POP Site",
        "Child Site connectivity",
        "Child Site Name",
        "Active",
        "Vendor",
        "Site Name S1",
        "Site Name S2",
        "Revise",
        "Protocol",
        "Status",
        "Model",
        "Bandwidth",
        "Planning Capacity",
        "Site S1 IP",
        "Site S2 IP",
        "Site S1 Port",
        "Site S2 Port",
    ]
    sheet.append(headers)

    for site_item, budget_item in rows:
        row = serialize_row(site_item, budget_item)
        sheet.append(
            [
                row["id"],
                row["sitea_id"],
                row["siteb_id"],
                row["link_id"],
                row["category_ne"],
                row["depth"],
                row["dependency"],
                row["pop_site"],
                row["child_site_connectivity"],
                row["child_site_name"],
                "Active" if row["is_active"] else "Inactive",
                row["budget_vendor"],
                row["budget_site_name_s1"],
                row["budget_site_name_s2"],
                row["budget_revise"],
                row["budget_protocol"],
                row["budget_status"],
                row["budget_model"],
                row["budget_bandwidth"],
                row["budget_planning_capacity"],
                row["budget_site_name_s1_ip"],
                row["budget_site_name_s2_ip"],
                row["budget_site_name_s1_port"],
                row["budget_site_name_s2_port"],
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=selected_site_connectivity.xlsx"
        },
    )
